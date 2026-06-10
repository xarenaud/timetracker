from flask import request, send_file  
from datetime import datetime  
import io

def register_export_routes(app, get_db, USE_PG, PLACEHOLDER, P, get_working_days):

    TARIF = 75.0  
    BORDEAUX_HEX = '7B1C2E'

    def get_dashboard_data(month):  
        if not month:  
            month = datetime.now().strftime('%Y-%m')  
        year_int, month_int = int(month.split('-')[0]), int(month.split('-')[1])  
from datetime import datetime
import io

def register_export_routes(app, get_db, USE_PG, PLACEHOLDER, P, get_working_days):

    TARIF = 75.0
    BORDEAUX_HEX = '7B1C2E'

    def get_dashboard_data(month):
        year_int, month_int = int(month.split('-')[0]), int(month.split('-')[1])
        working_days = get_working_days(year_int, month_int)
        working_weeks = round(working_days / 5, 2)

        conn = get_db()
        c = conn.cursor()

        c.execute("SELECT user_id, hourly_cost, vendable_hours FROM collaborator_settings")
        settings_raw = c.fetchall()
        settings = {}
        for r in settings_raw:
            if USE_PG:
                settings[r[0]] = {'hourly_cost': r[1], 'vendable_hours': r[2]}
            else:
                settings[r['user_id']] = {'hourly_cost': r['hourly_cost'], 'vendable_hours': r['vendable_hours']}

        if USE_PG:
            q = "SELECT te.user_id, u.username, SUM(te.duration_minutes) as total_minutes FROM time_entries te JOIN users u ON te.user_id = u.id WHERE TO_CHAR(CAST(te.start_time AS TIMESTAMP), %s) = %s GROUP BY te.user_id, u.username"
            c.execute(q, ('YYYY-MM', month))
        else:
            q = "SELECT te.user_id, u.username, SUM(te.duration_minutes) as total_minutes FROM time_entries te JOIN users u ON te.user_id = u.id WHERE strftime('%Y-%m', te.start_time) = ? GROUP BY te.user_id, u.username"
            c.execute(q, (month,))
        rows = c.fetchall()

        collab_stats = {}
        for r in rows:
            uid = r[0] if USE_PG else r['user_id']
            uname = r[1] if USE_PG else r['username']
            total_min = r[2] if USE_PG else r['total_minutes']
            total_h = (total_min or 0) / 60
            s = settings.get(uid, {'hourly_cost': 0, 'vendable_hours': 0})
            vendable_h_month = round(s['vendable_hours'] * working_weeks, 2)
            collab_stats[uid] = {
                'username': uname, 'total_hours': total_h,
                'vendable_hours_week': s['vendable_hours'],
                'vendable_hours': vendable_h_month,
                'hourly_cost': s['hourly_cost'],
                'ca_attendu': vendable_h_month * TARIF,
                'ca_realise': total_h * TARIF,
                'cout_total': total_h * s['hourly_cost'],
                'marge': (total_h * TARIF) - (total_h * s['hourly_cost']),
                'taux': round((total_h / vendable_h_month * 100), 1) if vendable_h_month > 0 else 0
            }

        if USE_PG:
            q2 = "SELECT cs.client_id, cl.name, cs.monthly_hours, COALESCE(SUM(te.duration_minutes), 0) as total_min FROM client_services cs JOIN clients cl ON cs.client_id = cl.id LEFT JOIN time_entries te ON te.client_id = cs.client_id AND TO_CHAR(CAST(te.start_time AS TIMESTAMP), %s) = %s GROUP BY cs.client_id, cl.name, cs.monthly_hours ORDER BY cl.name"
            c.execute(q2, ('YYYY-MM', month))
        else:
            q2 = "SELECT cs.client_id, cl.name, cs.monthly_hours, COALESCE(SUM(te.duration_minutes), 0) as total_min FROM client_services cs JOIN clients cl ON cs.client_id = cl.id LEFT JOIN time_entries te ON te.client_id = cs.client_id AND strftime('%Y-%m', te.start_time) = ? GROUP BY cs.client_id, cl.name, cs.monthly_hours ORDER BY cl.name"
            c.execute(q2, (month,))
        client_rows = c.fetchall()
        conn.close()

        client_stats = {}
        for r in client_rows:
            cid = r[0] if USE_PG else r['client_id']
            cname = r[1] if USE_PG else r['name']
            quota_h = r[2] if USE_PG else r['monthly_hours']
            total_min = r[3] if USE_PG else r['total_min']
            prested_h = (total_min or 0) / 60
            if cid not in client_stats:
                client_stats[cid] = {'name': cname, 'budget': 0, 'ca_realise': 0, 'hours_quota': 0, 'hours_prested': 0}
            client_stats[cid]['budget'] += quota_h * TARIF
            client_stats[cid]['ca_realise'] += prested_h * TARIF
            client_stats[cid]['hours_quota'] += quota_h
            client_stats[cid]['hours_prested'] += prested_h

        return collab_stats, client_stats, working_days, working_weeks

    @app.route('/dashboard/export/pdf')
    def export_pdf():
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt

        month = request.args.get('month', datetime.now().strftime('%Y-%m'))
        collab_stats, client_stats, working_days, working_weeks = get_dashboard_data(month)

        BORDEAUX_COLOR = colors.HexColor('#7B1C2E')
        GREEN_COLOR = colors.HexColor('#2e7d32')
        RED_COLOR = colors.HexColor('#c62828')

        def make_chart(labels, data1, data2, label1, label2, title):
            fig, ax = plt.subplots(figsize=(7, 3.5))
            x = range(len(labels))
            ax.bar([i-0.2 for i in x], data1, 0.4, label=label1, color='#7B1C2E', alpha=0.8)
            ax.bar([i+0.2 for i in x], data2, 0.4, label=label2, color='#2e7d32', alpha=0.8)
            ax.set_xticks(list(x))
            ax.set_xticklabels(labels, fontsize=9, rotation=15, ha='right')
            ax.set_ylabel('€', fontsize=9)
            ax.set_title(title, fontsize=11, fontweight='bold')
            ax.legend(fontsize=8)
            ax.grid(axis='y', alpha=0.3)
            plt.tight_layout()
            buf = io.BytesIO()
            plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
            plt.close()
            buf.seek(0)
            return buf

        buf_pdf = io.BytesIO()
        doc = SimpleDocTemplate(buf_pdf, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('t', parent=styles['Title'], textColor=BORDEAUX_COLOR, fontSize=18, spaceAfter=4)
        sub_style = ParagraphStyle('s', parent=styles['Normal'], textColor=colors.grey, fontSize=10, spaceAfter=12)
        section_style = ParagraphStyle('sec', parent=styles['Heading2'], textColor=BORDEAUX_COLOR, fontSize=13, spaceBefore=16, spaceAfter=8)
        footer_style = ParagraphStyle('f', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)

        story = []
        story.append(Paragraph("CX-Media TimeTracker", title_style))
        story.append(Paragraph(f"Rapport de rentabilité — {month} | {working_days} jours ouvrables | {working_weeks} semaines", sub_style))
        story.append(HRFlowable(width="100%", thickness=2, color=BORDEAUX_COLOR))
        story.append(Spacer(1, 0.4*cm))

        total_ca_attendu = sum(cs['ca_attendu'] for cs in collab_stats.values())
        total_ca_realise = sum(cs['ca_realise'] for cs in collab_stats.values())
        total_marge = sum(cs['marge'] for cs in collab_stats.values())
        total_cout = sum(cs['cout_total'] for cs in collab_stats.values())

        kpi_data = [
            ['CA Attendu', 'CA Réalisé', 'Marge Totale', 'Coûts Salariaux'],
            [f"{total_ca_attendu:.0f}€", f"{total_ca_realise:.0f}€", f"{total_marge:.0f}€", f"{total_cout:.0f}€"]
        ]
        kpi_t = Table(kpi_data, colWidths=[4.2*cm]*4)
        kpi_t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), BORDEAUX_COLOR), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,-1), 9),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,1), (-1,1), 14), ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#f5f5f5')),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#e0e0e0')),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e0e0e0')),
            ('TOPPADDING', (0,0), (-1,-1), 8), ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(kpi_t)
        story.append(Spacer(1, 0.5*cm))

        if collab_stats:
            story.append(Paragraph("Performance collaborateurs", section_style))
            chart1 = make_chart(
                [cs['username'] for cs in collab_stats.values()],
                [cs['ca_attendu'] for cs in collab_stats.values()],
                [cs['ca_realise'] for cs in collab_stats.values()],
                'CA Attendu', 'CA Réalisé', 'CA Attendu vs Réalisé par collaborateur'
            )
            story.append(Image(chart1, width=16*cm, height=8*cm))
            story.append(Spacer(1, 0.3*cm))

            collab_data = [['Collaborateur', 'H/sem', 'H/mois', 'H prestées', 'Taux', 'CA Attendu', 'CA Réalisé', 'Marge']]
            for cs in collab_stats.values():
                collab_data.append([cs['username'], f"{cs['vendable_hours_week']}h", f"{cs['vendable_hours']}h",
                    f"{cs['total_hours']:.1f}h", f"{cs['taux']}%",
                    f"{cs['ca_attendu']:.0f}€", f"{cs['ca_realise']:.0f}€", f"{cs['marge']:.0f}€"])
            t = Table(collab_data, colWidths=[3*cm,1.5*cm,1.8*cm,2*cm,1.5*cm,2.2*cm,2.2*cm,2*cm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), BORDEAUX_COLOR), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,-1), 8),
                ('ALIGN', (1,0), (-1,-1), 'CENTER'),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f9f9f9')]),
                ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#e0e0e0')),
                ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e0e0e0')),
                ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(t)

        if client_stats:
            story.append(Spacer(1, 0.5*cm))
            story.append(Paragraph("Rentabilité clients", section_style))
            chart2 = make_chart(
                [cs['name'] for cs in client_stats.values()],
                [cs['budget'] for cs in client_stats.values()],
                [cs['ca_realise'] for cs in client_stats.values()],
                'Budget client', 'CA Réalisé', 'Budget vs CA Réalisé par client'
            )
            story.append(Image(chart2, width=16*cm, height=8*cm))
            story.append(Spacer(1, 0.3*cm))

            client_data = [['Client', 'Budget', 'H. quota', 'H. prestées', 'Taux €/h', 'Statut']]
            for cs in client_stats.values():
                taux_h = cs['budget'] / cs['hours_prested'] if cs['hours_prested'] > 0 else TARIF
                statut = 'Excellent' if taux_h >= TARIF else ('Acceptable' if taux_h >= 35 else 'Danger')
                client_data.append([cs['name'], f"{cs['budget']:.0f}€", f"{cs['hours_quota']}h",
                    f"{cs['hours_prested']:.1f}h", f"{taux_h:.2f}€/h", statut])
            t2 = Table(client_data, colWidths=[4*cm,2.5*cm,2*cm,2.5*cm,2.5*cm,2.5*cm])
            t2.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), BORDEAUX_COLOR), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,-1), 8),
                ('ALIGN', (1,0), (-1,-1), 'CENTER'),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f9f9f9')]),
                ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#e0e0e0')),
                ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e0e0e0')),
                ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(t2)

        story.append(Spacer(1, 0.5*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e0e0e0')))
        story.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — CX-Media TimeTracker", footer_style))

        doc.build(story)
        buf_pdf.seek(0)
        return send_file(buf_pdf, mimetype='application/pdf', as_attachment=True, download_name=f'cx-media-rapport-{month}.pdf')

    @app.route('/dashboard/export/excel')
    def export_excel():
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, Reference

        month = request.args.get('month', datetime.now().strftime('%Y-%m'))
        collab_stats, client_stats, working_days, working_weeks = get_dashboard_data(month)

        wb = openpyxl.Workbook()
        WHITE = 'FFFFFF'

        header_font = Font(name='Calibri', bold=True, color=WHITE, size=11)
        header_fill = PatternFill(start_color=BORDEAUX_HEX, end_color=BORDEAUX_HEX, fill_type='solid')
        center = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0')
        )

        def style_header(ws, row, cols):
            for col in range(1, cols+1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = border

        def style_row(ws, row, cols, alt=False):
            fill = PatternFill(start_color='F9F9F9' if alt else WHITE, end_color='F9F9F9' if alt else WHITE, fill_type='solid')
            for col in range(1, cols+1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.alignment = center
                cell.border = border

        # Onglet Résumé
        ws1 = wb.active
        ws1.title = "Résumé"
        for col, w in zip(['A','B','C','D','E'], [25,18,18,18,18]):
            ws1.column_dimensions[col].width = w

        ws1['A1'] = f'CX-Media TimeTracker — Rapport {month}'
        ws1['A1'].font = Font(name='Calibri', bold=True, size=14, color=BORDEAUX_HEX)
        ws1['A2'] = f'{working_days} jours ouvrables | {working_weeks} semaines | Tarif: {TARIF}€/h'
        ws1['A2'].font = Font(name='Calibri', size=10, color='666666')
        ws1.merge_cells('A1:E1')
        ws1.merge_cells('A2:E2')
        ws1.append([])
        ws1.append(['KPI', 'CA Attendu', 'CA Réalisé', 'Marge Totale', 'Coûts Salariaux'])
        style_header(ws1, 4, 5)
        total_ca_attendu = sum(cs['ca_attendu'] for cs in collab_stats.values())
        total_ca_realise = sum(cs['ca_realise'] for cs in collab_stats.values())
        total_marge = sum(cs['marge'] for cs in collab_stats.values())
        total_cout = sum(cs['cout_total'] for cs in collab_stats.values())
        ws1.append(['Total', round(total_ca_attendu,0), round(total_ca_realise,0), round(total_marge,0), round(total_cout,0)])
        style_row(ws1, 5, 5)

        # Onglet Collaborateurs
        ws2 = wb.create_sheet("Collaborateurs")
        for col, w in zip(['A','B','C','D','E','F','G','H'], [20,12,14,12,10,14,14,14]):
            ws2.column_dimensions[col].width = w
        ws2.append(['Collaborateur', 'H/semaine', 'H vendables/mois', 'H prestées', 'Taux %', 'CA Attendu', 'CA Réalisé', 'Marge'])
        style_header(ws2, 1, 8)
        for i, cs in enumerate(collab_stats.values()):
            ws2.append([cs['username'], cs['vendable_hours_week'], cs['vendable_hours'],
                round(cs['total_hours'],2), cs['taux'], round(cs['ca_attendu'],0), round(cs['ca_realise'],0), round(cs['marge'],0)])
            style_row(ws2, i+2, 8, alt=(i%2==1))
        n = len(collab_stats)
        if n > 0:
            chart1 = BarChart()
            chart1.type = "col"
            chart1.title = "CA Attendu vs Réalisé"
            chart1.y_axis.title = "€"
            chart1.width = 18
            chart1.height = 12
            chart1.add_data(Reference(ws2, min_col=6, max_col=7, min_row=1, max_row=n+1), titles_from_data=True)
            chart1.set_categories(Reference(ws2, min_col=1, min_row=2, max_row=n+1))
            ws2.add_chart(chart1, f"A{n+4}")

        # Onglet Clients
        ws3 = wb.create_sheet("Clients")
        for col, w in zip(['A','B','C','D','E','F'], [22,14,12,14,16,14]):
            ws3.column_dimensions[col].width = w
        ws3.append(['Client', 'Budget', 'H. quota', 'H. prestées', 'Taux horaire €/h', 'Statut'])
        style_header(ws3, 1, 6)
        for i, cs in enumerate(client_stats.values()):
            taux_h = cs['budget'] / cs['hours_prested'] if cs['hours_prested'] > 0 else TARIF
            statut = 'Excellent' if taux_h >= TARIF else ('Acceptable' if taux_h >= 35 else 'Danger')
            ws3.append([cs['name'], round(cs['budget'],0), cs['hours_quota'], round(cs['hours_prested'],2), round(taux_h,2), statut])
            style_row(ws3, i+2, 6, alt=(i%2==1))
        nc = len(client_stats)
        if nc > 0:
            chart2 = BarChart()
            chart2.type = "col"
            chart2.title = "Budget vs CA Réalisé"
            chart2.y_axis.title = "€"
            chart2.width = 18
            chart2.height = 12
            chart2.add_data(Reference(ws3, min_col=2, max_col=2, min_row=1, max_row=nc+1), titles_from_data=True)
            chart2.add_data(Reference(ws3, min_col=4, max_col=4, min_row=1, max_row=nc+1), titles_from_data=True)
            chart2.set_categories(Reference(ws3, min_col=1, min_row=2, max_row=nc+1))
            ws3.add_chart(chart2, f"A{nc+4}")

        buf_excel = io.BytesIO()
        wb.save(buf_excel)
        buf_excel.seek(0)
        return send_file(buf_excel,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=f'cx-media-rapport-{month}.xlsx')
